"""
eBay Mini Fridge Scraper
------------------------
Scrapes eBay mini fridge listings, stores them in SQLite,
filters the cool branded ones via OpenAI, exports to Excel,
and visualizes all listings with Plotly.

Install deps:
    pip install playwright playwright-stealth parsel openai plotly python-dotenv openpyxl
    playwright install chromium

Set env:
    OPENAI_API_KEY=sk-...  (in .env file or export)
"""

import asyncio
import sqlite3
import json
import os
import re
import random
from pathlib import Path
from dotenv import load_dotenv
from playwright.async_api import async_playwright
try:
    from playwright_stealth import stealth_async
except Exception:
    stealth_async = None
from parsel import Selector
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# ── Config ─────────────────────────────────────────────────────────────────────
# All output files live in the same folder as this script.

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
HERE           = Path(__file__).parent
DB_PATH        = HERE / "fridges.db"
OUTPUT_HTML    = HERE / "fridge_results.html"
OUTPUT_EXCEL   = HERE / "fridge_listings.xlsx"
TARGET_URL     = (
    "https://www.ebay.com/sch/i.html"
    "?_nkw=mini+fridge&_sacat=0&_from=R40&LH_PrefLoc=5&rt=nc&_udhi=150"
)

# ── Database ───────────────────────────────────────────────────────────────────

def init_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS listings (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            title       TEXT,
            price       REAL,
            url         TEXT UNIQUE,
            image       TEXT,
            condition   TEXT,
            is_cool     INTEGER DEFAULT NULL,
            ai_reason   TEXT,
            scraped_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    print("✓ DB ready")
    return conn


def store_listings(conn: sqlite3.Connection, listings: list[dict]) -> int:
    inserted = 0
    for item in listings:
        try:
            conn.execute(
                """INSERT OR IGNORE INTO listings (title, price, url, image, condition)
                   VALUES (?, ?, ?, ?, ?)""",
                (item["title"], item["price"], item["url"], item["image"], item["condition"]),
            )
            if conn.execute("SELECT changes()").fetchone()[0]:
                inserted += 1
        except Exception as e:
            print(f"  DB insert error: {e}")
    conn.commit()
    return inserted

# ── Scraper ────────────────────────────────────────────────────────────────────

async def apply_basic_stealth(page):
    await page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] });
        Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
        window.chrome = window.chrome || { runtime: {} };
    """)


async def scrape_page(url: str) -> str:
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()
        if stealth_async:
            await stealth_async(page)
        else:
            await apply_basic_stealth(page)
            print("⚠ playwright-stealth unavailable; using basic stealth fallback")

        print("→ Navigating to eBay...")
        await page.goto(url, wait_until="domcontentloaded", timeout=45_000)
        try:
            await page.wait_for_selector("div.su-card-container, li.s-item", timeout=20_000)
        except Exception:
            print("⚠ Timed out waiting for listing cards; parsing whatever loaded")
        await asyncio.sleep(random.uniform(2, 4))

        for _ in range(4):
            await page.keyboard.press("End")
            await asyncio.sleep(random.uniform(0.8, 1.5))

        html = await page.content()
        await browser.close()
        print("✓ Page fetched")
        return html

# ── Parser ─────────────────────────────────────────────────────────────────────

def _extract_price(raw: str) -> float | None:
    if not raw:
        return None
    match = re.search(r"[\d,]+\.?\d*", raw.replace(",", ""))
    return float(match.group()) if match else None


def parse_listings(html: str) -> list[dict]:
    sel     = Selector(text=html)
    results = []

    for item in sel.css(".s-item"):
        title = item.css(".s-item__title::text").get("").strip()
        if not title or title.lower() == "shop on ebay":
            continue
        price_raw = (
            item.css(".s-item__price::text").get("")
            or item.css(".s-item__price .ITALIC::text").get("")
        )
        url   = item.css(".s-item__link::attr(href)").get("")
        image = (
            item.css(".s-item__image-img::attr(src)").get("")
            or item.css(".s-item__image-img::attr(data-src)").get("")
        )
        condition = (
            item.css(".SECONDARY_INFO::text").get("")
            or item.css(".s-item__condition::text").get("")
            or "Unknown"
        ).strip()
        results.append({"title": title, "price": _extract_price(price_raw),
                         "url": url, "image": image, "condition": condition})

    if results:
        return results

    # Fallback: new eBay card layout (2026)
    for card in sel.css("div.su-card-container"):
        title = card.css(".s-card__title .su-styled-text::text").get("").strip()
        if not title:
            title = " ".join(t.strip() for t in card.css(".s-card__title ::text").getall() if t.strip())
        if not title or title.lower() == "shop on ebay":
            continue
        price_raw = card.css(".s-card__price::text").get("")
        url = card.css("a.s-card__link::attr(href), a[href*='/itm/']::attr(href)").get("")
        if "/itm/123456" in url:
            continue
        image = (
            card.css("img.s-card__image-img::attr(src)").get("")
            or card.css("img::attr(src)").get("")
            or card.css("img::attr(data-src)").get("")
        )
        condition = (
            card.css(".s-card__subtitle .su-styled-text::text").get("")
            or card.css(".s-card__subtitle::text").get("")
            or "Unknown"
        ).strip()
        results.append({"title": title, "price": _extract_price(price_raw),
                         "url": url, "image": image, "condition": condition})

    return results

# ── AI Filter ──────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """
You are filtering eBay mini fridge listings to find the ones that are visually 'cool'
or collectible — specifically:

COOL = branded with Monster Energy, Red Bull, Coca-Cola, Pepsi, Dr Pepper,
       Sprite, Heineken, Budweiser, gaming brands (Xbox, PlayStation, Razer),
       retro/vintage designs, neon aesthetics, or any statement piece fridge
       that's more than a plain white box.

NOT COOL = generic white/black mini fridges, plain hotel-style, no branding.

Respond ONLY with a raw JSON array. No markdown fences. No commentary.
Format strictly:
[{"id": 1, "is_cool": true, "reason": "short reason"}, ...]
""".strip()


def filter_with_ai(conn: sqlite3.Connection):
    if not OPENAI_API_KEY:
        print("⚠ No OPENAI_API_KEY — skipping AI filter")
        return

    client = OpenAI(api_key=OPENAI_API_KEY)
    rows = conn.execute("SELECT id, title FROM listings WHERE is_cool IS NULL").fetchall()

    if not rows:
        print("✓ All listings already classified")
        return

    print(f"→ Running AI filter on {len(rows)} listings...")
    BATCH = 30
    for i in range(0, len(rows), BATCH):
        batch   = rows[i : i + BATCH]
        payload = json.dumps([{"id": r[0], "title": r[1]} for r in batch])
        try:
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": f"Classify these listings:\n{payload}"},
                ],
                temperature=0,
                max_tokens=800,
            )
            raw = re.sub(r"```json|```", "", resp.choices[0].message.content.strip()).strip()
            for r in json.loads(raw):
                conn.execute(
                    "UPDATE listings SET is_cool = ?, ai_reason = ? WHERE id = ?",
                    (1 if r.get("is_cool") else 0, r.get("reason", ""), r["id"]),
                )
            conn.commit()
            print(f"  Batch {i//BATCH + 1} done ({len(batch)} items)")
        except (json.JSONDecodeError, KeyError) as e:
            print(f"  ⚠ Parse error on batch {i//BATCH + 1}: {e}")
        except Exception as e:
            print(f"  ⚠ API error: {e}")

# ── Excel Export ───────────────────────────────────────────────────────────────

def export_excel(conn: sqlite3.Connection):
    all_rows  = conn.execute(
        "SELECT id, title, price, condition, is_cool, ai_reason, url, scraped_at FROM listings ORDER BY id"
    ).fetchall()
    cool_rows = [r for r in all_rows if r[4] == 1]

    wb      = openpyxl.Workbook()
    headers = ["ID", "Title", "Price ($)", "Condition", "Cool?", "AI Reason", "URL", "Scraped At"]
    col_widths = [6, 60, 12, 18, 10, 40, 80, 22]

    def style_header(cell, fill_color):
        cell.fill      = PatternFill("solid", fgColor=fill_color)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Sheet 1: All Listings ──────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Listings"
    cool_fill = PatternFill("solid", fgColor="FFE0B2")

    for col, h in enumerate(headers, 1):
        style_header(ws_all.cell(row=1, column=col, value=h), "1F4E79")

    for row_idx, row in enumerate(all_rows, 2):
        id_, title, price, condition, is_cool, ai_reason, url, scraped_at = row
        values = [
            id_, title,
            round(price, 2) if price else None,
            condition,
            "Yes" if is_cool == 1 else ("No" if is_cool == 0 else "Unclassified"),
            ai_reason or "", url, scraped_at,
        ]
        for col, val in enumerate(values, 1):
            cell = ws_all.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=(col in (2, 6)), vertical="top")
            if is_cool == 1:
                cell.fill = cool_fill

    for col, width in enumerate(col_widths, 1):
        ws_all.column_dimensions[get_column_letter(col)].width = width
    ws_all.freeze_panes  = "A2"
    ws_all.auto_filter.ref = ws_all.dimensions

    # ── Sheet 2: Cool Ones ─────────────────────────────────────────────────────
    ws_cool = wb.create_sheet("Cool Ones")

    for col, h in enumerate(headers, 1):
        style_header(ws_cool.cell(row=1, column=col, value=h), "BF360C")

    for row_idx, row in enumerate(cool_rows, 2):
        id_, title, price, condition, is_cool, ai_reason, url, scraped_at = row
        values = [id_, title, round(price, 2) if price else None,
                  condition, "Yes", ai_reason or "", url, scraped_at]
        for col, val in enumerate(values, 1):
            cell = ws_cool.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=(col in (2, 6)), vertical="top")

    for col, width in enumerate(col_widths, 1):
        ws_cool.column_dimensions[get_column_letter(col)].width = width
    ws_cool.freeze_panes   = "A2"
    ws_cool.auto_filter.ref = ws_cool.dimensions

    wb.save(OUTPUT_EXCEL)
    print(f"✓ Excel saved → {OUTPUT_EXCEL}  ({len(all_rows)} total, {len(cool_rows)} cool)")

# ── Visualise ──────────────────────────────────────────────────────────────────

def visualize(conn: sqlite3.Connection):
    all_rows   = conn.execute(
        "SELECT id, title, price, condition, is_cool, ai_reason, url FROM listings ORDER BY price"
    ).fetchall()
    cool_rows  = [r for r in all_rows if r[4] == 1 and r[2] is not None]
    other_rows = [r for r in all_rows if r[4] != 1 and r[2] is not None]
    cool_prices  = [r[2] for r in cool_rows]
    other_prices = [r[2] for r in other_rows]

    def row_html(r):
        id_, title, price, condition, is_cool, reason, url = r
        price_str  = f"${price:.2f}" if price else "N/A"
        cool_badge = (
            '<span style="color:#e63946;font-weight:bold">🔥 Cool</span>' if is_cool == 1
            else ('<span style="color:#aaa">Regular</span>' if is_cool == 0
                  else '<span style="color:#888">—</span>')
        )
        bg = ' style="background:#2a1a1a"' if is_cool == 1 else ""
        title_cell = (f'<a href="{url}" target="_blank" style="color:#ccc;text-decoration:none">{title}</a>'
                      if url else title)
        return (f"<tr{bg}><td>{id_}</td><td>{title_cell}</td>"
                f"<td style='text-align:right'>{price_str}</td>"
                f"<td>{condition}</td><td>{cool_badge}</td>"
                f"<td style='color:#888;font-size:0.85em'>{reason or ''}</td></tr>")

    table_rows = "".join(row_html(r) for r in all_rows)

    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=("Price Distribution — Cool 🔥 vs Regular", "Cool Fridges by Price"),
        row_heights=[0.55, 0.45],
        vertical_spacing=0.12,
    )
    fig.add_trace(go.Histogram(x=other_prices, name="Regular",
                               marker_color="#555", opacity=0.6, nbinsx=20), row=1, col=1)
    fig.add_trace(go.Histogram(x=cool_prices,  name="Cool 🔥",
                               marker_color="#e63946", opacity=0.9, nbinsx=20), row=1, col=1)

    if cool_rows:
        short_titles = [(t[:40] + "…" if len(t) > 40 else t) for _, t, *_ in cool_rows]
        fig.add_trace(
            go.Scatter(
                x=cool_prices, y=list(range(len(cool_prices))),
                mode="markers+text", name="Cool Listings",
                text=short_titles, textposition="middle right",
                marker=dict(color="#e63946", size=14, symbol="star"),
                hovertext=[f"<b>${r[2]:.2f}</b><br>{r[1]}<br><i>{r[5]}</i>" for r in cool_rows],
                hoverinfo="text",
            ),
            row=2, col=1,
        )

    fig.update_layout(title="eBay Mini Fridge Scrape Results", template="plotly_dark",
                      barmode="overlay", height=850, showlegend=True,
                      font=dict(family="monospace"))
    fig.update_xaxes(title_text="Price (USD)", row=1, col=1)
    fig.update_yaxes(title_text="Count",       row=1, col=1)
    fig.update_xaxes(title_text="Price (USD)", row=2, col=1)
    fig.update_yaxes(visible=False,            row=2, col=1)

    chart_div = fig.to_html(full_html=False, include_plotlyjs="cdn")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>eBay Mini Fridge Results</title>
<style>
  body {{ background:#111; color:#ddd; font-family:monospace; margin:0; padding:20px; }}
  h1   {{ color:#e63946; }}
  h2   {{ color:#aaa; border-bottom:1px solid #333; padding-bottom:6px; }}
  .stats {{ display:flex; gap:30px; margin-bottom:20px; }}
  .stat  {{ background:#1e1e1e; border-radius:8px; padding:14px 24px; }}
  .stat span {{ display:block; font-size:2em; font-weight:bold; color:#e63946; }}
  table {{ width:100%; border-collapse:collapse; font-size:0.88em; }}
  th    {{ background:#1F4E79; color:#fff; padding:8px 10px; text-align:left; position:sticky; top:0; }}
  td    {{ padding:7px 10px; border-bottom:1px solid #222; vertical-align:top; }}
  tr:hover td {{ background:#1a2530 !important; }}
  .tbl-wrap {{ max-height:600px; overflow-y:auto; border:1px solid #333; border-radius:6px; margin-top:10px; }}
</style>
</head>
<body>
<h1>🧊 eBay Mini Fridge Scrape Results</h1>
<div class="stats">
  <div class="stat"><span>{len(all_rows)}</span>Total Scraped</div>
  <div class="stat"><span>{len(cool_rows)}</span>Cool Fridges 🔥</div>
  <div class="stat"><span>{len(other_rows)}</span>Regular Fridges</div>
</div>
{chart_div}
<h2>All Scraped Listings ({len(all_rows)})</h2>
<div class="tbl-wrap">
<table>
  <thead>
    <tr><th>#</th><th>Title</th><th>Price</th><th>Condition</th><th>Cool?</th><th>AI Reason</th></tr>
  </thead>
  <tbody>{table_rows}</tbody>
</table>
</div>
</body>
</html>"""

    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"✓ HTML saved → {OUTPUT_HTML}")

# ── Console Summary ────────────────────────────────────────────────────────────

def print_summary(conn: sqlite3.Connection):
    total    = conn.execute("SELECT COUNT(*) FROM listings").fetchone()[0]
    cool_cnt = conn.execute("SELECT COUNT(*) FROM listings WHERE is_cool = 1").fetchone()[0]
    all_rows = conn.execute(
        "SELECT title, price, condition, is_cool, ai_reason FROM listings ORDER BY is_cool DESC, price"
    ).fetchall()

    print(f"\n{'═'*70}")
    print(f"  TOTAL SCRAPED : {total}")
    print(f"  COOL FRIDGES  : {cool_cnt}")
    print(f"{'═'*70}\n")
    print(f"  {'PRICE':>8}  {'':^4}  TITLE")
    print(f"  {'─'*8}  {'─'*4}  {'─'*55}")
    for title, price, condition, is_cool, reason in all_rows:
        price_str = f"${price:.2f}" if price else "  N/A "
        flag      = "🔥" if is_cool == 1 else "  "
        print(f"  {price_str:>8}  {flag}  {title[:60]}")
        if is_cool == 1 and reason:
            print(f"  {'':>8}       ↳ {reason}")
    print()

# ── Entry Point ────────────────────────────────────────────────────────────────

async def main():
    print("🧊 eBay Mini Fridge Scraper\n")
    print(f"📁 Folder: {HERE.resolve()}\n")

    conn = init_db()

    html     = await scrape_page(TARGET_URL)
    listings = parse_listings(html)
    print(f"✓ Parsed {len(listings)} raw listings")

    inserted = store_listings(conn, listings)
    print(f"✓ {inserted} new listings saved to DB")

    filter_with_ai(conn)
    export_excel(conn)
    visualize(conn)
    print_summary(conn)

    conn.close()
    print(f"📁 All files in: {HERE.resolve()}")


if __name__ == "__main__":
    asyncio.run(main())
